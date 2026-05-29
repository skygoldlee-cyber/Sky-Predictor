"""성능 분석기

거래 성능을 심층 분석하는 리포트를 생성합니다.

Usage:
    from prediction.performance_analyzer import PerformanceAnalyzer
    from trading.state import TradeRecord
    
    analyzer = PerformanceAnalyzer()
    report = analyzer.generate_report(trades)
    analyzer.print_report(report)
    
    # 시각화
    analyzer.plot_equity_curve(trades)
    analyzer.plot_performance_by_time(trades)
    
    # 보고서 저장
    analyzer.save_report_to_excel(report, "performance_report.xlsx")
"""

import logging
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field
from pathlib import Path
import pandas as pd

try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
_logger = logging.getLogger(__name__)


@dataclass
class PerformanceReport:
    """성능 분석 리포트."""
    basic_stats: Dict[str, Any]
    time_analysis: Dict[str, Any]
    day_of_week_analysis: Dict[str, Any]
    confidence_analysis: Dict[str, Any]
    exit_reason_analysis: Dict[str, Any]
    holding_time_analysis: Dict[str, Any]
    risk_metrics: Dict[str, Any]
    excursion_analysis: Dict[str, Any]
    slot_analysis: Dict[str, Any] = field(default_factory=dict)
    sizing_analysis: Dict[str, Any] = field(default_factory=dict)
    advanced_metrics: Dict[str, Any] = field(default_factory=dict)


class PerformanceAnalyzer:
    """성능 분석기."""
    
    def __init__(self):
        """초기화."""
        pass
    
    def generate_report(self, trades: List[Any]) -> PerformanceReport:
        """상세 성능 리포트 생성.
        
        Args:
            trades: 거래 리스트 (TradeRecord 객체)
        
        Returns:
            성능 분석 리포트
        """
        if not trades:
            return PerformanceReport(
                basic_stats={},
                time_analysis={},
                day_of_week_analysis={},
                confidence_analysis={},
                exit_reason_analysis={},
                holding_time_analysis={},
                risk_metrics={},
                excursion_analysis={}
            )
        
        # 데이터프레임 변환
        df = self._trades_to_dataframe(trades)
        
        return PerformanceReport(
            basic_stats=self._calculate_basic_stats(df),
            time_analysis=self._analyze_by_time(df),
            day_of_week_analysis=self._analyze_by_day(df),
            confidence_analysis=self._analyze_by_confidence(df),
            exit_reason_analysis=self._analyze_by_exit_reason(df),
            holding_time_analysis=self._analyze_holding_time(df),
            risk_metrics=self._calculate_risk_metrics(df),
            excursion_analysis=self._analyze_excursions(df),
            slot_analysis=self._analyze_by_slot(df),
            sizing_analysis=self._analyze_sizing(df),
            advanced_metrics=self._calculate_advanced_metrics(df)
        )
    
    def _trades_to_dataframe(self, trades: List[Any]) -> pd.DataFrame:
        """거래 리스트를 데이터프레임으로 변환."""
        data = []
        for trade in trades:
            # TradeRecord 객체인지 확인
            if hasattr(trade, 'slot'):
                # TradeRecord
                data.append({
                    "entry_time": trade.entry_time,
                    "exit_time": trade.close_time,
                    "entry_price": trade.entry_price,
                    "exit_price": trade.close_price,
                    "action": "BUY" if trade.side.value == "LONG" else "SELL",
                    "exit_reason": trade.close_reason.value if trade.close_reason else None,
                    "profit": trade.pnl_pt,
                    "profit_pct": trade.pnl_pt / trade.entry_price * 100 if trade.entry_price > 0 else 0,
                    "bars_held": trade.hold_minutes,
                    "slot": trade.slot.value,
                    "confidence": trade.entry_confidence,
                    "position_size": getattr(trade, 'position_size', 0.0),
                    "capital_used": getattr(trade, 'capital_used', 0.0),
                    "risk_amount": getattr(trade, 'risk_amount', 0.0),
                    "sizing_method": getattr(trade, 'sizing_method', ''),
                })
            else:
                # Backtest Trade 객체
                data.append({
                    "entry_time": trade.entry_time,
                    "exit_time": trade.exit_time,
                    "entry_price": trade.entry_price,
                    "exit_price": trade.exit_price,
                    "action": trade.action,
                    "exit_reason": trade.exit_reason,
                    "profit": trade.profit,
                    "profit_pct": trade.profit_pct,
                    "bars_held": trade.bars_held,
                    "slot": "UNKNOWN",
                    "confidence": "MEDIUM",
                    "position_size": 0.0,
                    "capital_used": 0.0,
                    "risk_amount": 0.0,
                    "sizing_method": "",
                })
        return pd.DataFrame(data)
    
    def _calculate_basic_stats(self, df: pd.DataFrame) -> Dict[str, Any]:
        """기본 통계 계산."""
        total_trades = len(df)
        win_trades = len(df[df["profit"] > 0])
        loss_trades = len(df[df["profit"] <= 0])
        win_rate = win_trades / total_trades if total_trades > 0 else 0
        
        total_profit = df["profit"].sum()
        avg_profit = df["profit"].mean()
        avg_profit_pct = df["profit_pct"].mean()
        
        max_profit = df["profit"].max()
        max_loss = df["profit"].min()
        
        avg_holding_time = df["bars_held"].mean()
        
        return {
            "total_trades": total_trades,
            "win_trades": win_trades,
            "loss_trades": loss_trades,
            "win_rate": win_rate,
            "total_profit": total_profit,
            "avg_profit": avg_profit,
            "avg_profit_pct": avg_profit_pct,
            "max_profit": max_profit,
            "max_loss": max_loss,
            "avg_holding_time": avg_holding_time
        }
    
    def _analyze_by_time(self, df: pd.DataFrame) -> Dict[str, Any]:
        """시간대별 분석."""
        df["entry_hour"] = df["entry_time"].dt.hour
        
        time_slots = {
            "morning (09:00-11:30)": df[(df["entry_hour"] >= 9) & (df["entry_hour"] < 12)],
            "afternoon (13:00-14:30)": df[(df["entry_hour"] >= 13) & (df["entry_hour"] < 15)],
            "close (14:30-15:30)": df[(df["entry_hour"] >= 14) & (df["entry_hour"] < 16)]
        }
        
        analysis = {}
        for slot, slot_df in time_slots.items():
            if len(slot_df) > 0:
                analysis[slot] = {
                    "count": len(slot_df),
                    "win_rate": len(slot_df[slot_df["profit"] > 0]) / len(slot_df),
                    "avg_profit": slot_df["profit"].mean(),
                    "total_profit": slot_df["profit"].sum()
                }
        
        return analysis
    
    def _analyze_by_day(self, df: pd.DataFrame) -> Dict[str, Any]:
        """요일별 분석."""
        df["entry_day"] = df["entry_time"].dt.day_name()
        
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        analysis = {}
        
        for day in days:
            day_df = df[df["entry_day"] == day]
            if len(day_df) > 0:
                analysis[day] = {
                    "count": len(day_df),
                    "win_rate": len(day_df[day_df["profit"] > 0]) / len(day_df),
                    "avg_profit": day_df["profit"].mean(),
                    "total_profit": day_df["profit"].sum()
                }
        
        return analysis
    
    def _analyze_by_confidence(self, df: pd.DataFrame) -> Dict[str, Any]:
        """신뢰도별 분석."""
        # Trade 객체에 confidence 필드가 있는지 확인 필요
        # 현재 Trade 데이터클래스에는 confidence 필드가 없으므로
        # 추후 추가 필요
        
        return {
            "note": "신뢰도별 분석은 Trade 객체에 confidence 필드 추가 후 구현"
        }
    
    def _analyze_by_exit_reason(self, df: pd.DataFrame) -> Dict[str, Any]:
        """청산 사유별 분석."""
        analysis = {}
        
        for reason in df["exit_reason"].unique():
            if pd.notna(reason):
                reason_df = df[df["exit_reason"] == reason]
                analysis[reason] = {
                    "count": len(reason_df),
                    "win_rate": len(reason_df[reason_df["profit"] > 0]) / len(reason_df),
                    "avg_profit": reason_df["profit"].mean(),
                    "total_profit": reason_df["profit"].sum()
                }
        
        return analysis
    
    def _analyze_holding_time(self, df: pd.DataFrame) -> Dict[str, Any]:
        """보유 기간 분석."""
        df["holding_time_category"] = pd.cut(
            df["bars_held"],
            bins=[0, 5, 15, 30, float('inf')],
            labels=["short (0-5)", "medium (5-15)", "long (15-30)", "very_long (30+)"]
        )
        
        analysis = {}
        for category in df["holding_time_category"].cat.categories:
            cat_df = df[df["holding_time_category"] == category]
            if len(cat_df) > 0:
                analysis[str(category)] = {
                    "count": len(cat_df),
                    "win_rate": len(cat_df[cat_df["profit"] > 0]) / len(cat_df),
                    "avg_profit": cat_df["profit"].mean()
                }
        
        return analysis
    
    def _calculate_risk_metrics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """리스크 메트릭 계산."""
        profits = df["profit"].values
        
        # MDD 계산
        cumulative = pd.Series(profits).cumsum()
        running_max = cumulative.expanding().max()
        drawdown = (cumulative - running_max) / running_max
        max_drawdown = drawdown.min()
        
        # Sharpe Ratio (단순화)
        if len(profits) > 1:
            returns = pd.Series(profits).pct_change().dropna()
            sharpe_ratio = returns.mean() / returns.std() * (252 ** 0.5) if returns.std() > 0 else 0
        else:
            sharpe_ratio = 0
        
        # Sortino Ratio
        if len(profits) > 1:
            downside_returns = returns[returns < 0]
            sortino_ratio = returns.mean() / downside_returns.std() * (252 ** 0.5) if len(downside_returns) > 0 and downside_returns.std() > 0 else 0
        else:
            sortino_ratio = 0
        
        # Profit Factor
        total_profit = profits[profits > 0].sum()
        total_loss = abs(profits[profits < 0].sum())
        profit_factor = total_profit / total_loss if total_loss > 0 else 0
        
        return {
            "max_drawdown": max_drawdown,
            "max_drawdown_pct": max_drawdown * 100,
            "sharpe_ratio": sharpe_ratio,
            "sortino_ratio": sortino_ratio,
            "profit_factor": profit_factor
        }
    
    def _analyze_excursions(self, df: pd.DataFrame) -> Dict[str, Any]:
        """최대 호재/악재 분석."""
        # 현재 Trade 데이터클래스에는 호재/악재 필드가 없으므로
        # BacktestPositionState의 필드를 참조하거나 추후 추가 필요
        
        return {
            "note": "호재/악재 분석은 Trade 객체에 max_favorable_excursion, max_adverse_excursion 필드 추가 후 구현"
        }
    
    def _analyze_by_slot(self, df: pd.DataFrame) -> Dict[str, Any]:
        """슬롯별 분석."""
        analysis = {}
        
        for slot in df["slot"].unique():
            if pd.notna(slot):
                slot_df = df[df["slot"] == slot]
                if len(slot_df) > 0:
                    analysis[slot] = {
                        "count": len(slot_df),
                        "win_rate": len(slot_df[slot_df["profit"] > 0]) / len(slot_df),
                        "avg_profit": slot_df["profit"].mean(),
                        "total_profit": slot_df["profit"].sum(),
                        "avg_holding_time": slot_df["bars_held"].mean()
                    }
        
        return analysis
    
    def _analyze_sizing(self, df: pd.DataFrame) -> Dict[str, Any]:
        """포지션 사이징 분석."""
        # 사이징 방법별 분석
        sizing_method_analysis = {}
        for method in df["sizing_method"].unique():
            if pd.notna(method) and method:
                method_df = df[df["sizing_method"] == method]
                if len(method_df) > 0:
                    sizing_method_analysis[method] = {
                        "count": len(method_df),
                        "win_rate": len(method_df[method_df["profit"] > 0]) / len(method_df),
                        "avg_profit": method_df["profit"].mean(),
                        "total_profit": method_df["profit"].sum(),
                        "avg_position_size": method_df["position_size"].mean(),
                        "avg_risk_pct": method_df["risk_amount"].mean() / method_df["capital_used"].mean() if method_df["capital_used"].mean() > 0 else 0
                    }
        
        # 리스크 노출 분석
        total_risk = df["risk_amount"].sum()
        avg_risk_per_trade = df["risk_amount"].mean()
        
        return {
            "sizing_method_analysis": sizing_method_analysis,
            "total_risk_exposure": total_risk,
            "avg_risk_per_trade": avg_risk_per_trade,
            "avg_position_size": df["position_size"].mean(),
            "avg_capital_used": df["capital_used"].mean()
        }
    
    def _calculate_advanced_metrics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """고급 성과 지표 계산."""
        profits = df["profit"].values
        
        # Calmar Ratio (연간 수익률 / 최대 낙폭)
        if len(profits) > 1:
            cumulative = pd.Series(profits).cumsum()
            running_max = cumulative.expanding().max()
            drawdown = (cumulative - running_max) / running_max
            max_drawdown = abs(drawdown.min())
            
            # 연간 수익률 추정 (거래 기간 기반)
            if len(df) > 1:
                days = (df["exit_time"].max() - df["entry_time"].min()).days
                if days > 0:
                    annual_return = cumulative.iloc[-1] / days * 365
                    calmar_ratio = annual_return / max_drawdown if max_drawdown > 0 else 0
                else:
                    calmar_ratio = 0
            else:
                calmar_ratio = 0
        else:
            calmar_ratio = 0
        
        # Win/Loss Ratio
        avg_win = profits[profits > 0].mean() if len(profits[profits > 0]) > 0 else 0
        avg_loss = abs(profits[profits < 0].mean()) if len(profits[profits < 0]) > 0 else 0
        win_loss_ratio = avg_win / avg_loss if avg_loss > 0 else 0
        
        # Expectancy (기대값)
        win_rate = len(profits[profits > 0]) / len(profits) if len(profits) > 0 else 0
        expectancy = (win_rate * avg_win) - ((1 - win_rate) * avg_loss)
        
        # Risk of Ruin (파산 위험도)
        # 단순화: 연속 손실 확률 기반
        if avg_loss > 0:
            # 초기 자본 대비 평균 손실
            initial_capital = 1000000  # 가정
            avg_loss_pct = avg_loss / initial_capital
            risk_of_ruin = ((1 - win_rate) / win_rate) ** (initial_capital / avg_loss) if win_rate > 0 else 1
            risk_of_ruin = min(risk_of_ruin, 1.0)  # 최대 1.0
        else:
            risk_of_ruin = 0
        
        # Recovery Factor (총 수익 / 최대 낙폭)
        if len(profits) > 1:
            cumulative = pd.Series(profits).cumsum()
            running_max = cumulative.expanding().max()
            drawdown = (cumulative - running_max) / running_max
            max_drawdown = abs(drawdown.min())
            recovery_factor = cumulative.iloc[-1] / max_drawdown if max_drawdown > 0 else 0
        else:
            recovery_factor = 0
        
        return {
            "calmar_ratio": calmar_ratio,
            "win_loss_ratio": win_loss_ratio,
            "expectancy": expectancy,
            "risk_of_ruin": risk_of_ruin,
            "recovery_factor": recovery_factor,
            "avg_win": avg_win,
            "avg_loss": avg_loss
        }
    
    def print_report(self, report: PerformanceReport):
        """성능 리포트 출력."""
        print("\n" + "="*60)
        print("성능 분석 리포트")
        print("="*60)
        
        print("\n[기본 통계]")
        for key, value in report.basic_stats.items():
            if isinstance(value, float):
                print(f"{key}: {value:.2f}")
            else:
                print(f"{key}: {value}")
        
        print("\n[시간대별 분석]")
        for slot, stats in report.time_analysis.items():
            print(f"{slot}:")
            print(f"  거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}, 총 수익: {stats['total_profit']:,.0f}")
        
        print("\n[요일별 분석]")
        for day, stats in report.day_of_week_analysis.items():
            print(f"{day}:")
            print(f"  거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}, 총 수익: {stats['total_profit']:,.0f}")
        
        print("\n[슬롯별 분석]")
        for slot, stats in report.slot_analysis.items():
            print(f"{slot}:")
            print(f"  거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}, 총 수익: {stats['total_profit']:,.2f}pt")
        
        print("\n[신뢰도별 분석]")
        for confidence, stats in report.confidence_analysis.items():
            if isinstance(stats, dict) and 'count' in stats:
                print(f"{confidence}:")
                print(f"  거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}, 총 수익: {stats['total_profit']:,.0f}")
            else:
                print(f"{confidence}: {stats}")
        
        print("\n[청산 사유별 분석]")
        for reason, stats in report.exit_reason_analysis.items():
            print(f"{reason}:")
            print(f"  거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}, 총 수익: {stats['total_profit']:,.0f}")
        
        print("\n[보유 기간 분석]")
        for category, stats in report.holding_time_analysis.items():
            print(f"{category}:")
            print(f"  거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}, 평균 수익: {stats['avg_profit']:,.0f}")
        
        print("\n[리스크 메트릭]")
        for key, value in report.risk_metrics.items():
            if isinstance(value, float):
                print(f"{key}: {value:.4f}")
            else:
                print(f"{key}: {value}")
        
        print("\n[고급 성과 지표]")
        for key, value in report.advanced_metrics.items():
            if isinstance(value, float):
                print(f"{key}: {value:.4f}")
            else:
                print(f"{key}: {value}")
        
        print("\n[포지션 사이징 분석]")
        if report.sizing_analysis:
            print(f"총 리스크 노출: {report.sizing_analysis.get('total_risk_exposure', 0):,.0f}")
            print(f"평균 거래당 리스크: {report.sizing_analysis.get('avg_risk_per_trade', 0):,.0f}")
            print(f"평균 포지션 사이즈: {report.sizing_analysis.get('avg_position_size', 0):.2f}")
            print(f"평균 사용 자본: {report.sizing_analysis.get('avg_capital_used', 0):,.0f}")
            
            sizing_methods = report.sizing_analysis.get('sizing_method_analysis', {})
            if sizing_methods:
                print("\n사이징 방법별:")
                for method, stats in sizing_methods.items():
                    print(f"  {method}:")
                    print(f"    거래 수: {stats['count']}, 승률: {stats['win_rate']:.2%}")
                    print(f"    평균 사이즈: {stats['avg_position_size']:.2f}")
        
        print("="*60)
    
    def plot_equity_curve(self, trades: List[Any], save_path: Optional[Path] = None):
        """자본 곡선 시각화.
        
        Args:
            trades: 거래 리스트
            save_path: 저장 경로 (None이면 표시만)
        """
        if not MATPLOTLIB_AVAILABLE:
            _logger.warning("[PERF] matplotlib이 설치되지 않아 시각화를 건너뜁니다.")
            return
        
        df = self._trades_to_dataframe(trades)
        if len(df) == 0:
            _logger.warning("[PERF] 거래 데이터가 없습니다.")
            return
        
        # 누적 수익 계산
        df['cumulative_profit'] = df['profit'].cumsum()
        
        plt.figure(figsize=(12, 6))
        plt.plot(df['exit_time'], df['cumulative_profit'], linewidth=2)
        plt.axhline(y=0, color='r', linestyle='--', alpha=0.5)
        plt.xlabel('날짜')
        plt.ylabel('누적 수익 (pt)')
        plt.title('자본 곡선')
        plt.grid(True, alpha=0.3)
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            _logger.info("[PERF] 자본 곡선 저장: %s", save_path)
        else:
            plt.show()
        
        plt.close()
    
    def plot_performance_by_time(self, trades: List[Any], save_path: Optional[Path] = None):
        """시간대별 성과 시각화.
        
        Args:
            trades: 거래 리스트
            save_path: 저장 경로 (None이면 표시만)
        """
        if not MATPLOTLIB_AVAILABLE:
            _logger.warning("[PERF] matplotlib이 설치되지 않아 시각화를 건너뜁니다.")
            return
        
        df = self._trades_to_dataframe(trades)
        if len(df) == 0:
            _logger.warning("[PERF] 거래 데이터가 없습니다.")
            return
        
        df['entry_hour'] = df['entry_time'].dt.hour
        
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        
        # 거래 수
        ax = axes[0, 0]
        hourly_counts = df.groupby('entry_hour').size()
        ax.bar(hourly_counts.index, hourly_counts.values)
        ax.set_xlabel('시간')
        ax.set_ylabel('거래 수')
        ax.set_title('시간대별 거래 수')
        ax.grid(True, alpha=0.3)
        
        # 승률
        ax = axes[0, 1]
        hourly_win_rate = df.groupby('entry_hour')['profit'].apply(lambda x: (x > 0).mean())
        ax.bar(hourly_win_rate.index, hourly_win_rate.values)
        ax.axhline(y=0.5, color='r', linestyle='--', alpha=0.5)
        ax.set_xlabel('시간')
        ax.set_ylabel('승률')
        ax.set_title('시간대별 승률')
        ax.grid(True, alpha=0.3)
        
        # 평균 수익
        ax = axes[1, 0]
        hourly_profit = df.groupby('entry_hour')['profit'].mean()
        colors = ['green' if p > 0 else 'red' for p in hourly_profit.values]
        ax.bar(hourly_profit.index, hourly_profit.values, color=colors)
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        ax.set_xlabel('시간')
        ax.set_ylabel('평균 수익 (pt)')
        ax.set_title('시간대별 평균 수익')
        ax.grid(True, alpha=0.3)
        
        # 총 수익
        ax = axes[1, 1]
        hourly_total_profit = df.groupby('entry_hour')['profit'].sum()
        colors = ['green' if p > 0 else 'red' for p in hourly_total_profit.values]
        ax.bar(hourly_total_profit.index, hourly_total_profit.values, color=colors)
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        ax.set_xlabel('시간')
        ax.set_ylabel('총 수익 (pt)')
        ax.set_title('시간대별 총 수익')
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            _logger.info("[PERF] 시간대별 성과 저장: %s", save_path)
        else:
            plt.show()
        
        plt.close()
    
    def save_report_to_excel(self, report: PerformanceReport, output_path: Path):
        """성과 리포트를 Excel로 저장.
        
        Args:
            report: 성과 리포트
            output_path: 출력 경로
        """
        if not OPENPYXL_AVAILABLE:
            _logger.warning("[PERF] openpyxl이 설치되지 않아 Excel 저장을 건너뜁니다.")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "성과 분석 리포트"
        
        # 헤더 스타일
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        row = 1
        
        # 기본 통계
        ws.cell(row=row, column=1, value="기본 통계").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = header_alignment
        row += 1
        
        for key, value in report.basic_stats.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # 리스크 메트릭
        ws.cell(row=row, column=1, value="리스크 메트릭").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = header_alignment
        row += 1
        
        for key, value in report.risk_metrics.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # 고급 성과 지표
        ws.cell(row=row, column=1, value="고급 성과 지표").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = header_alignment
        row += 1
        
        for key, value in report.advanced_metrics.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # 슬롯별 분석
        ws.cell(row=row, column=1, value="슬롯별 분석").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = header_alignment
        row += 1
        
        for slot, stats in report.slot_analysis.items():
            ws.cell(row=row, column=1, value=slot)
            ws.cell(row=row, column=2, value=f"거래 수: {stats['count']}")
            ws.cell(row=row, column=3, value=f"승률: {stats['win_rate']:.2%}")
            ws.cell(row=row, column=4, value=f"총 수익: {stats['total_profit']:.2f}pt")
            row += 1
        
        row += 1
        
        # 시간대별 분석
        ws.cell(row=row, column=1, value="시간대별 분석").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = header_alignment
        row += 1
        
        for slot, stats in report.time_analysis.items():
            ws.cell(row=row, column=1, value=slot)
            ws.cell(row=row, column=2, value=f"거래 수: {stats['count']}")
            ws.cell(row=row, column=3, value=f"승률: {stats['win_rate']:.2%}")
            ws.cell(row=row, column=4, value=f"총 수익: {stats['total_profit']:.0f}")
            row += 1
        
        row += 1
        
        # 포지션 사이징 분석
        if report.sizing_analysis:
            ws.cell(row=row, column=1, value="포지션 사이징 분석").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).alignment = header_alignment
            row += 1
            
            ws.cell(row=row, column=1, value="총 리스크 노출")
            ws.cell(row=row, column=2, value=report.sizing_analysis.get('total_risk_exposure', 0))
            row += 1
            
            ws.cell(row=row, column=1, value="평균 거래당 리스크")
            ws.cell(row=row, column=2, value=report.sizing_analysis.get('avg_risk_per_trade', 0))
            row += 1
            
            ws.cell(row=row, column=1, value="평균 포지션 사이즈")
            ws.cell(row=row, column=2, value=report.sizing_analysis.get('avg_position_size', 0))
            row += 1
            
            sizing_methods = report.sizing_analysis.get('sizing_method_analysis', {})
            if sizing_methods:
                row += 1
                ws.cell(row=row, column=1, value="사이징 방법별").font = Font(bold=True)
                row += 1
                
                for method, stats in sizing_methods.items():
                    ws.cell(row=row, column=1, value=method)
                    ws.cell(row=row, column=2, value=f"거래 수: {stats['count']}")
                    ws.cell(row=row, column=3, value=f"승률: {stats['win_rate']:.2%}")
                    row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        wb.save(output_path)
        _logger.info("[PERF] Excel 보고서 저장: %s", output_path)
